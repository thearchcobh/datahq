from __future__ import annotations

import os
import re
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any

import requests
from openpyxl import load_workbook

from .database import get_client

DEFAULT_WORKBOOK_URL = (
    "https://www.dropbox.com/scl/fi/w1dlfox32i6tnid2ryq4j/"
    "Actual_and_ProjectedHours2023_2026.xlsx"
    "?rlkey=flnn2glff51q399wno58mp77b&dl=1"
)
SOURCE_FILE = "Actual_and_ProjectedHours2023_2026.xlsx"
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
TIME_RANGE_RE = re.compile(r"^\s*(\d{1,2}):(\d{2})\s*[-–—]\s*(\d{1,2}):(\d{2})\s*$")


def _parse_week_start(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    raise ValueError(f"Unrecognised Week Starting value: {value!r}")


def _parse_schedule(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    upper = text.upper()
    # Handles Closed, CLOSED XMAS, and historical typos such as Clsosed.
    if "CLOS" in upper or "CLS" in upper:
        return {
            "is_scheduled_open": False,
            "open_time": None,
            "close_time": None,
            "close_next_day": False,
            "schedule_text": text,
        }

    match = TIME_RANGE_RE.match(text)
    if not match:
        raise ValueError(f"Unrecognised opening-hours cell: {text!r}")

    sh, sm, eh, em = map(int, match.groups())
    if not (0 <= sh <= 23 and 0 <= eh <= 23 and 0 <= sm <= 59 and 0 <= em <= 59):
        raise ValueError(f"Invalid time range: {text!r}")

    start = time(sh, sm)
    end = time(eh, em)
    return {
        "is_scheduled_open": True,
        "open_time": start.isoformat(timespec="minutes"),
        "close_time": end.isoformat(timespec="minutes"),
        "close_next_day": end <= start,
        "schedule_text": text,
    }


def _find_schedule_sheet(workbook):
    required = {"Week Starting", *DAY_NAMES}
    for sheet in workbook.worksheets:
        for row_number in range(1, min(sheet.max_row, 20) + 1):
            values = [sheet.cell(row_number, col).value for col in range(1, min(sheet.max_column, 20) + 1)]
            normalized = {str(v).strip() for v in values if v is not None}
            if required.issubset(normalized):
                columns = {
                    str(sheet.cell(row_number, col).value).strip(): col
                    for col in range(1, sheet.max_column + 1)
                    if sheet.cell(row_number, col).value is not None
                }
                return sheet, row_number, columns
    raise RuntimeError("Could not find opening-hours header row in workbook")


def read_workbook_rows(content: bytes) -> tuple[list[dict[str, Any]], date, date]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    sheet, header_row, columns = _find_schedule_sheet(workbook)

    rows: list[dict[str, Any]] = []
    seen: set[date] = set()
    min_date: date | None = None
    max_date: date | None = None

    for row_number in range(header_row + 1, sheet.max_row + 1):
        raw_week_start = sheet.cell(row_number, columns["Week Starting"]).value
        if raw_week_start is None or str(raw_week_start).strip() == "":
            continue
        try:
            week_start = _parse_week_start(raw_week_start)
        except ValueError:
            continue

        min_date = week_start if min_date is None else min(min_date, week_start)
        max_date = week_start + timedelta(days=6) if max_date is None else max(max_date, week_start + timedelta(days=6))

        for offset, day_name in enumerate(DAY_NAMES):
            trading_date = week_start + timedelta(days=offset)
            parsed = _parse_schedule(sheet.cell(row_number, columns[day_name]).value)
            if parsed is None:
                # Blank means unknown/not supplied, not automatically closed.
                continue
            if trading_date in seen:
                raise RuntimeError(f"Duplicate opening-hours date in workbook: {trading_date}")
            seen.add(trading_date)
            rows.append(
                {
                    "trading_date": trading_date.isoformat(),
                    "source_week_start": week_start.isoformat(),
                    "source_file": SOURCE_FILE,
                    "source_precision": "exact_minutes",
                    **parsed,
                }
            )

    if not rows or min_date is None or max_date is None:
        raise RuntimeError("Workbook did not contain any usable opening-hours rows")
    return rows, min_date, max_date


def sync_opening_hours() -> None:
    url = os.getenv("OPENING_HOURS_WORKBOOK_URL") or DEFAULT_WORKBOOK_URL
    response = requests.get(url, timeout=60, allow_redirects=True)
    response.raise_for_status()
    rows, min_date, max_date = read_workbook_rows(response.content)

    client = get_client()
    # Replace the workbook's covered range so removed/blank future dates do not linger as stale schedule rows.
    client.table("analytics_opening_hours_daily").delete().gte(
        "trading_date", min_date.isoformat()
    ).lte("trading_date", max_date.isoformat()).execute()

    for start in range(0, len(rows), 500):
        client.table("analytics_opening_hours_daily").upsert(
            rows[start : start + 500], on_conflict="trading_date"
        ).execute()

    print(
        f"Opening hours synced from {SOURCE_FILE}: {len(rows)} dated schedule rows "
        f"covering {min_date} through {max_date}."
    )


if __name__ == "__main__":
    sync_opening_hours()
